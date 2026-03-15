"""
Create Excel reports from aggregated_accuracy_curves.csv.

For each (generator, judge, fill) combination, creates an Excel file with:
- Pivot table: method x occlusion_level with mean_acc values (methods as rows)
- Excel chart created from the pivot table data
"""

import pandas as pd
from pathlib import Path
import logging
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def create_excel_with_chart(
    df: pd.DataFrame,
    generator: str,
    judge: str,
    fill: str,
    output_path: Path
) -> int:
    """
    Create Excel file with pivot table and method-comparison chart.
    """
    # Create pivot table: method as rows, occlusion_level as columns
    pivot_df = df.pivot_table(
        index='attribution_method',
        columns='occlusion_level',
        values='mean_accuracy',
        aggfunc='mean'
    )
    
    # Reset index to make method a column
    pivot_df = pivot_df.reset_index()
    
    # Sort occlusion levels (columns)
    # Filter out 'method' column, convert rest to float for sorting, then sort
    occlusion_cols = [c for c in pivot_df.columns if c != 'attribution_method']
    # Sort columns based on their numeric value (handles string '0.1' vs float 0.1)
    try:
        occlusion_cols.sort(key=lambda x: float(x))
    except ValueError:
        occlusion_cols.sort() # Fallback to string sort
        
    # Reorder DataFrame: Method column first, then sorted occlusion levels
    pivot_df = pivot_df[['attribution_method'] + occlusion_cols]
    
    num_rows = len(pivot_df) + 1  # +1 for header
    num_cols = len(pivot_df.columns)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write pivot table to sheet
        pivot_df.to_excel(writer, sheet_name='Data', index=False)
        
        # Get workbook and worksheet
        worksheet = writer.sheets['Data']
        
        # --- Formatting ---
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # --- Chart Creation ---
        chart = LineChart()
        chart.title = f"Accuracy Degradation\nGenerator: {generator} | Judge: {judge} | Fill: {fill}"
        chart.style = 10
        chart.y_axis.title = 'Top-1 Accuracy'
        chart.x_axis.title = 'Occlusion Level (%)'
        chart.height = 15
        chart.width = 25  # Made it slightly wider
        
        # DATA REFERENCE
        # min_col=1 includes the "Method" column so it can be used for Legend Titles
        # min_row=2 starts from the first data row (skipping header)
        data = Reference(worksheet, min_col=1, min_row=2, max_row=num_rows, max_col=num_cols)
        
        # Add data to chart
        # from_rows=True: Each ROW is a line (Series) on the graph
        # titles_from_data=True: Use the first column (Method) as the series name
        chart.add_data(data, from_rows=True, titles_from_data=True)
        
        # CATEGORY AXIS (X-Axis)
        # min_col=2 starts from the first occlusion level (skipping "Method" header)
        cats = Reference(worksheet, min_col=2, min_row=1, max_col=num_cols)
        chart.set_categories(cats)
        
        # Set y-axis limits (0 to 1 accuracy)
        chart.y_axis.scaling.min = 0.0
        chart.y_axis.scaling.max = 1.05
        
        # Position legend at top-right
        chart.legend.position = 'tr'  # 'tr' = top-right
        
        # Position chart below the table
        chart_start_row = num_rows + 3
        chart_cell = f'A{chart_start_row}'
        worksheet.add_chart(chart, chart_cell)
    
    logging.info(f"Created Excel file: {output_path}")
    return num_rows


def main():
    """Main function to process CSV and create Excel reports."""
    # Paths
    base_dir = Path(__file__).parent.parent
    csv_path = base_dir / "results" / "analysis" / "aggregated_accuracy_curves.csv"
    output_dir = base_dir / "results"
    
    if not csv_path.exists():
        logging.error(f"CSV file not found: {csv_path}")
        return
    
    # Load CSV
    logging.info(f"Loading CSV: {csv_path}")
    df = pd.read_csv(csv_path)
    
    # Clean column names (remove spaces)
    df.columns = df.columns.str.strip()
    
    # Check and rename columns to standard names
    if 'DATA' in df.columns:
        df = df.rename(columns={'DATA': 'dataset'})
    elif 'data' in df.columns:
        df = df.rename(columns={'data': 'dataset'})
        
    # Standardize column names
    df = df.rename(columns={col: col.lower() for col in df.columns})

    # Ensure occlusion_level is numeric for proper sorting/plotting
    if 'occlusion_level' in df.columns:
        df['occlusion_level'] = pd.to_numeric(df['occlusion_level'], errors='coerce')
    
    # Ensure we have the required columns
    required_cols = ['generating_model', 'attribution_method', 'judging_model', 'fill_strategy', 'occlusion_level', 'mean_accuracy']
    missing_cols = [col for col in required_cols if col not in df.columns]
    
    if missing_cols:
        logging.error(f"Required columns not found: {missing_cols}")
        logging.error(f"Available columns: {df.columns.tolist()}")
        return
    
    # Group by (generator, judge, fill)
    group_cols = ['generating_model', 'judging_model', 'fill_strategy']
    
    # Filter out any rows with NaN in grouping columns
    df = df.dropna(subset=group_cols)
    
    combinations = df.groupby(group_cols)
    
    logging.info(f"Found {len(combinations)} unique combinations")
    
    # Create output directory
    excel_output_dir = output_dir / "excel_reports"
    excel_output_dir.mkdir(exist_ok=True, parents=True)
    
    # Process each combination
    for (generator, judge, fill), group_df in combinations:
        # Create filename
        filename = f"{generator}_{judge}_{fill}.xlsx"
        # Clean filename (remove invalid characters)
        filename = filename.replace('/', '_').replace('\\', '_').replace(':', '_')
        output_path = excel_output_dir / filename
        
        logging.info(f"Processing: Generator={generator}, Judge={judge}, Fill={fill}")
        
        try:
            # Create Excel with pivot table and chart
            create_excel_with_chart(group_df, generator, judge, fill, output_path)
            
        except Exception as e:
            logging.error(f"Error processing {generator}_{judge}_{fill}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    logging.info(f"Excel reports created in: {excel_output_dir}")


if __name__ == "__main__":
    main()